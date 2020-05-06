import os,sys,errno
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
# 
def silentremove(filename):
    try:
        os.remove(filename)
        print(filename + " File not transferred and will be removed")
    except OSError as e:                #this would be "except OSError, e:" before Python 2.6
        if e.errno!=errno.ENOENT:       #errno.ENOENT = no such file or directory
            raise                       #re-raise exception if a different error occured.

def Sort(file_to_shift,folder_to_put):
    for subdir, dirs, files in os.walk(rootdir):
        for file in files:
            base_file, ext = os.path.splitext(file)
            filepath = subdir + os.sep + file
            if filepath.endswith(file_to_shift):
                full_newpath = subdir  + os.sep + folder_to_put + os.sep + file
                # print(full_newpath)
                if (subdir == currDir + os.sep + folder_to_put):
                    pass
                elif (subdir != subdir + os.sep +  folder_to_put):
                    try:
                        os.rename(filepath, full_newpath)
                    except FileExistsError as error:
                        silentremove(filepath)
                        pass

def findstartRow(x,ReferencedNosofRow):
    for row in ws.iter_rows(min_row=1,max_col=1,max_row = ReferencedNosofRow):
                for cell in row:      
                    if cell.value ==str(x):
                        x_start_new=cell.col_idx
                        y_start_new=cell.row
    return x_start_new, y_start_new

def findendRow(y_start,ReferencedNosofRow):
    y_term_counter = 0
    for row in ws.iter_rows(min_row=y_start,max_col=1,max_row = ReferencedNosofRow):
        for cell in row:
            cell_row = cell.row
            if cell.value == None:
                y_term_counter = y_term_counter + 1
                if y_term_counter >1:
                    break
                x_term=cell.col_idx
                y_term=cell.row
    return x_term,y_term

def createFolder(folder_path):
    try:
        os.mkdir(folder_path)
    except OSError:
        print ("Creation of the directory %s failed, the file already exist" % folder_path)    

createFolder("XLSX_converted FOLDER - Compilation")
createFolder("AGS TO XLSX - Compilation")

AGS_file_counter = 0

currDir = os.path.dirname(os.path.realpath(__file__))
rootdir = os.path.abspath(os.path.join(currDir, '..'))
if rootdir not in sys.path: # add parent dir to paths
    sys.path.append(rootdir)

for subdir, dirs, files in os.walk(currDir):
    for file in files:
        base_file, ext = os.path.splitext(file)
        filepath = subdir + os.sep + file
        
        if filepath.endswith(".ags") or filepath.endswith(".AGS"):
            input_file = subdir + os.sep + base_file + ".ags"
            
            output_file = currDir + os.sep + base_file + "-" + str(AGS_file_counter) + "-BASE.xlsx"
            
            wb_create1 = openpyxl.Workbook()
            source = wb_create1.active
            target = wb_create1.copy_worksheet(source)
            ws_create1 = wb_create1.worksheets[0]
            ws_create1.title = "BASE FILE"
            ws_create2 = wb_create1.worksheets[1]
            ws_create2.title = "COORDINATE"
            with open(input_file) as data:
                reader = csv.reader(data, delimiter=',')
                for row in reader:
                    ws_create1.append(row)
                    
            wb_create1.save(output_file)
            wb = load_workbook(output_file)
            sheet_ranges = wb["BASE FILE"]
            ws = wb.active
            ws1 = wb.worksheets[1]
            #print(ws1)
            ws.title = "SPT"
            ReferencedNosofCol = ws.max_column
            ReferencedNosofRow = ws.max_row

            x_start_GEOL,y_start_GEOL = findstartRow("**GEOL",ReferencedNosofRow)
            x_start_ISPT, y_start_ISPT = findstartRow("**ISPT",ReferencedNosofRow)
            x_start_Hole, y_start_HOLE = findstartRow("**HOLE",ReferencedNosofRow)
            x_term_GEOL,y_term_GEOL = findendRow(y_start_GEOL,ReferencedNosofRow)
            x_term_HOLE,y_term_HOLE = findendRow(y_start_HOLE,ReferencedNosofRow)
            x_term_ISPT,y_term_ISPT = findendRow(y_start_ISPT,ReferencedNosofRow)

            for row in ws.iter_cols(min_row=1,max_col=6,max_row = ReferencedNosofRow):
                for cell in row:
                    if cell.row>=y_start_GEOL and cell.row <y_term_GEOL and cell.col_idx==1:
                        holeID = ws.cell(row = cell.row-y_start_GEOL+1,column = cell.col_idx+ReferencedNosofCol,value=cell.value)
                        #print(holeID.coordinate,holeID.value)
                    if cell.row>=y_start_GEOL and cell.row <y_term_GEOL and cell.col_idx==2:
                        GEOL_TOP=ws.cell(row = cell.row-y_start_GEOL+1,column = cell.col_idx+ReferencedNosofCol+1,value=cell.value)
                        #print(GEOL_TOP.coordinate,GEOL_TOP.value)
                    if cell.row>=y_start_GEOL and cell.row <y_term_GEOL and cell.col_idx==3:
                        GEOL_BASE=ws.cell(row = cell.row-y_start_GEOL+1,column = cell.col_idx+ReferencedNosofCol+1,value=cell.value)
                        #print(GEOL_BASE.coordinate,GEOL_BASE.value)
                    if cell.row>=y_start_GEOL and cell.row <y_term_GEOL and cell.col_idx==6:
                        GEOL_GEOL=ws.cell(row = cell.row-y_start_GEOL+1,column = cell.col_idx+ReferencedNosofCol+1,value=cell.value)
                        #print(GEOL_GEOL.coordinate,GEOL_GEOL.value)

            max_row_holeID = holeID.row
            max_column_holeID = holeID.col_idx

            holeID_col_ref = holeID.col_idx

            col_holeID = ws['W']
            SPT_counter = 0


            # In[14]:

            SPT_counter = 0
            for col_holeID in ws.iter_cols(min_row=1, max_col=max_column_holeID, max_row=max_row_holeID):
                        for holeID in col_holeID:
                            if holeID.col_idx == int(max_column_holeID):
                                for col in ws.iter_cols(min_row=y_start_ISPT+4, max_col=4, max_row=y_term_ISPT-1):
                                    #print(col)
                                    for cell in col:
                                        if cell.col_idx==1:
                                            ISPT_TOP=ws.cell(row=cell.row,column=cell.col_idx+1)
                                            #print("ISPT_TOP is at ",ISPT_TOP.coordinate,ISPT_TOP.value)
                                            ISPT_MAIN=ws.cell(row=cell.row,column=cell.col_idx+3)
                                            #print("ISPT_MAIN is at ",ISPT_MAIN.coordinate,ISPT_MAIN.value)
                                            if cell.value == holeID.value:
                                                GEOL_TOP=ws.cell(row=holeID.row,column=holeID.col_idx+2)
                                                GEOL_BASE=ws.cell(row=holeID.row,column=holeID.col_idx+3)
                                                GEOL_GEOL=ws.cell(row=holeID.row,column=holeID.col_idx+6)
                                                #print(GEOL_TOP.value,GEOL_BASE.value,ISPT_TOP.value)
                                                if float(ISPT_TOP.value)>=float(GEOL_TOP.value):
                                                    if float(ISPT_TOP.value)<float(GEOL_BASE.value):
                                                        SPT_counter = int(SPT_counter) + 1
                                                        ISPT_MAIN_SPT = ws.cell(row=holeID.row,column=holeID.col_idx+6+SPT_counter,value = float(ISPT_MAIN.value))
                                                        #print(holeID.col_idx,ISPT_MAIN_SPT.col_idx,GEOL_TOP.value,GEOL_BASE.value,ISPT_TOP.value,"ISPT_TOP is within the range of ",GEOL_TOP.value," and ", GEOL_BASE.value,"ISPT MAIN is at ",ISPT_MAIN_SPT.coordinate,ISPT_MAIN_SPT.value)                                

                                                else:
                                                    #print("ISPT_TOP is not within the range of ",GEOL_TOP.value," and ", GEOL_BASE.value)
                                                    SPT_counter = 0
                                    else:
                                        SPT_counter = 0

                                for col_hole in ws.iter_cols(min_row=y_start_HOLE+3, max_col=5, max_row=y_term_HOLE-1):
                                    #print(col_hole)
                                    for cell_hole in col_hole:
                                        if cell_hole.col_idx==1:
                                            HOLE_GL=ws.cell(row=cell_hole.row,column=cell_hole.col_idx+4)
                                            #print(HOLE_GL.coordinate,HOLE_GL.value)
                                            if cell_hole.value == holeID.value:
                                                    SPT_counter = SPT_counter + 1
                                                    HOLE_GL_new = ws.cell(row=holeID.row,column=holeID.col_idx+1,value = float(HOLE_GL.value))
                                                    #print(holeID.coordinate,holeID.value,"SOIL TOP LEVEL is at ",HOLE_GL_new.coordinate,HOLE_GL_new.value)                             
                                        else:
                                            SPT_counter = 0                                    

            ws.delete_cols(1,ReferencedNosofCol)

            ReferencedNosofCol_new = ws.max_column
            ReferencedNosofRow_new = ws.max_row            
            
            for row in ws.iter_rows(min_row=4,max_col=ReferencedNosofCol_new,max_row = ReferencedNosofRow_new):
                for cell_HOLE_GL in row:
                    if cell_HOLE_GL.col_idx==2:
                        if cell_HOLE_GL.value != None:
                            GEOL_TOP=ws.cell(row = cell_HOLE_GL.row,column = cell_HOLE_GL.col_idx+1)
                            GEOL_BASE=ws.cell(row = cell_HOLE_GL.row,column = cell_HOLE_GL.col_idx+2)
                            HOLE_GL_TOP = ws.cell(row = cell_HOLE_GL.row, column = cell_HOLE_GL.col_idx + 3, value = float(cell_HOLE_GL.value) - float(GEOL_TOP.value) - 100)
                            HOLE_GL_BASE = ws.cell(row = cell_HOLE_GL.row, column = cell_HOLE_GL.col_idx + 4, value = float(cell_HOLE_GL.value) - float(GEOL_BASE.value) - 100)
                        elif cell_HOLE_GL.value == None:
                            break
                        
            output_file_converted = currDir + os.sep + base_file + "-" + str(AGS_file_counter) + "_converted.xlsx"
            wb.save(output_file_converted)
            AGS_file_counter = AGS_file_counter + 1
                       
file_counter_base = 0

currDir = os.path.dirname(os.path.realpath(__file__))
rootdir = os.path.abspath(os.path.join(currDir, '..'))
if rootdir not in sys.path: # add parent dir to paths
    sys.path.append(rootdir)

for subdir, dirs, files in os.walk(rootdir):
    for file in files:
        base_file, ext = os.path.splitext(file)
        #print os.path.join(subdir, file)
        filepath = subdir + os.sep + file
        combined_file = 'Compilation of AGS.xlsx'
        if filepath.endswith("_converted.xlsx"):
            file_counter_base = file_counter_base + 1
            base_file = filepath
            wb_base = load_workbook(base_file)
            if file_counter_base == 1:
                wb_combined = load_workbook(combined_file)
                ws_base=wb_base.active
                ws_combined = wb_combined.active

                RefNosofCol_ws_base = ws_base.max_column
                RefNosofRow_ws_base = ws_base.max_row

                RefNosofCol_ws_combined = ws_combined.max_column
                RefNosofRow_ws_combined = ws_combined.max_row

                for row_base in ws_base.iter_rows(min_row=4, max_col=RefNosofCol_ws_base, max_row=RefNosofRow_ws_base):
                    #print(RefNosofCol_ws_base,row_base)
                    for cell_base in row_base:
                        #print(type(cell_base),cell_base,cell_base.value)
                        cell_to_be_copied = cell_base
                        cell_to_be_copied.value = cell_base.value
                        #print(type(cell_to_be_copied),cell_to_be_copied,cell_to_be_copied.value)
                        cell_combined = ws_combined.cell(row = RefNosofRow_ws_combined+1,column = cell_to_be_copied.col_idx,value = cell_to_be_copied.value)
                        #print(cell_combined,cell_combined.value)
                    RefNosofRow_ws_combined = ws_combined.max_row

                RefNosofCol_ws_combined = ws_combined.max_column
                RefNosofRow_ws_combined = ws_combined.max_row
                wb_combined.save('Compilation of AGS - combined.xlsx')
                
            elif file_counter_base > 1:
                combined_file = 'Compilation of AGS - combined.xlsx'
                wb_combined = load_workbook(combined_file)

                ws_base=wb_base.active
                ws_combined = wb_combined.active

                RefNosofCol_ws_base = ws_base.max_column
                RefNosofRow_ws_base = ws_base.max_row

                RefNosofCol_ws_combined = ws_combined.max_column
                RefNosofRow_ws_combined = ws_combined.max_row
        
                for row_base in ws_base.iter_rows(min_row=4, max_col=RefNosofCol_ws_base, max_row=RefNosofRow_ws_base):
                    #print(RefNosofCol_ws_base,row_base)
                    for cell_base in row_base:
                        #print(type(cell_base),cell_base,cell_base.value)
                        cell_to_be_copied = cell_base
                        cell_to_be_copied.value = cell_base.value
                        #print(type(cell_to_be_copied),cell_to_be_copied,cell_to_be_copied.value)
                        cell_combined = ws_combined.cell(row = RefNosofRow_ws_combined+1,column = cell_to_be_copied.col_idx,value = cell_to_be_copied.value)
                        #print(cell_combined,cell_combined.value)
                    RefNosofRow_ws_combined = ws_combined.max_row

                RefNosofCol_ws_combined = ws_combined.max_column
                RefNosofRow_ws_combined = ws_combined.max_row
                wb_combined.save('Compilation of AGS - combined.xlsx')

Sort("_converted.xlsx","XLSX_converted FOLDER - Compilation")
Sort("-BASE.xlsx","AGS TO XLSX - Compilation")
print("Conversion is Completed!")
print("You can now proceed to use the files")



